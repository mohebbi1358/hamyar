from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
from django.contrib import messages
from notification.models import NotificationGroup, NotificationCoupon
from donation.models import Donation
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from notification.models import NotificationGroup, NotificationCoupon
from donation.models import Donation
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from martyrs.models import Martyr
from eternals.models import Eternals
from wallet.models import Wallet, WalletTransaction
from .utils import process_wallet_donation
from donation.models import Donation
import random







@login_required
def donate(request):
    causes = DonationCause.objects.filter(is_active=True)

    martyr = None
    count = Martyr.objects.count()
    if count > 0:
        martyr = Martyr.objects.all()[random.randint(0, count - 1)]

    if request.method == 'POST':
        amount_str = request.POST.get('amount', '').replace(',', '')
        cause_id = request.POST.get('cause')
        pay_method = request.POST.get('pay_method')

        try:
            amount = int(amount_str)
            if amount < 1000:
                messages.error(request, "حداقل مبلغ ۱۰۰۰ تومان است.")
                return redirect('donation:donate')
        except ValueError:
            messages.error(request, "مبلغ وارد شده معتبر نیست.")
            return redirect('donation:donate')

        if not cause_id:
            messages.error(request, "لطفاً دلیل صدقه را انتخاب کنید.")
            return redirect('donation:donate')

        cause_obj = get_object_or_404(DonationCause, id=cause_id)

        if pay_method == 'wallet':
            success, error_msg, donation = process_wallet_donation(
                request.user,
                amount,
                cause_obj
            )

            if not success:
                messages.error(request, error_msg)
                return redirect('donation:donate')

            messages.success(request, "پرداخت از طریق کیف پول با موفقیت انجام شد.")
            return redirect('donations:donation_detail', pk=donation.id)

        elif pay_method == 'gateway':
            donation = Donation.objects.create(
                amount=amount,
                cause=cause_obj,
                status='PENDING',
                user=request.user,
                pay_method='gateway'
            )

            
            # ذخیره آدرس بازگشت در سشن (می‌تونی در صورت نیاز تغییر بدی)
            request.session["payment_return_url"] = reverse("donation:donate")
            
            return redirect('donation:fake_gateway', donation_id=donation.id)

        else:
            messages.error(request, "روش پرداخت نامعتبر است.")
            return redirect('donation:donate')

    suggested_amounts = [10000, 25000, 40000, 60000, 100000]
    return render(request, 'donation/donate.html', {
        'causes': causes,
        'suggested_amounts': suggested_amounts,
        'martyr': martyr,
        'eternal': None,
    })





@login_required
def donate_for_eternal(request, eternal_id):
    causes = DonationCause.objects.filter(is_active=True)
    eternal = get_object_or_404(Eternals, pk=eternal_id)

    next_url = request.GET.get("next")
    if next_url:
        request.session["payment_return_url"] = next_url

    if request.method == 'POST':
        amount_str = request.POST.get('amount', '').replace(',', '')
        pay_method = request.POST.get('pay_method')
        cause_id = request.POST.get('cause')

        try:
            amount = int(amount_str)
            if amount < 1000:
                messages.error(request, "حداقل مبلغ ۱۰۰۰ تومان است.")
                return redirect('donation:donate_for_eternal', eternal_id=eternal_id)
        except ValueError:
            messages.error(request, "مبلغ وارد شده معتبر نیست.")
            return redirect('donation:donate_for_eternal', eternal_id=eternal_id)

        if not cause_id:
            messages.error(request, "لطفاً دلیل صدقه را انتخاب کنید.")
            return redirect('donation:donate_for_eternal', eternal_id=eternal_id)

        cause_obj = get_object_or_404(DonationCause, id=cause_id)

        if pay_method == 'wallet':
            success, error_msg, donation = process_wallet_donation(
                request.user,
                amount,
                cause_obj,
                martyr=None,
                eternal=eternal
            )
            if not success:
                messages.error(request, error_msg)
                return redirect('donation:donate_for_eternal', eternal_id=eternal_id)

            messages.success(request, "پرداخت از طریق کیف پول با موفقیت انجام شد.")
            return redirect('donations:donation_detail', pk=donation.id)

        elif pay_method == 'gateway':
            donation = Donation.objects.create(
                amount=amount,
                cause=cause_obj,
                status='PENDING',
                user=request.user,
                eternal=eternal,
                pay_method='gateway'
            )

            # ذخیره آدرس بازگشت در سشن
            request.session["payment_return_url"] = reverse("eternals:detail", kwargs={"pk": eternal.id})

            return redirect('donation:fake_gateway', donation_id=donation.id)

        else:
            messages.error(request, "روش پرداخت نامعتبر است.")
            return redirect('donation:donate_for_eternal', eternal_id=eternal_id)

    suggested_amounts = [10000, 25000, 40000, 60000, 100000]
    return render(request, 'donation/donate.html', {
        'causes': causes,
        'eternal': eternal,
        'martyr': None,
        'suggested_amounts': suggested_amounts,
    })








@login_required
def donate_for_martyr(request, martyr_id):
    causes = DonationCause.objects.filter(is_active=True)
    martyr = get_object_or_404(Martyr, pk=martyr_id)

    if request.method == 'POST':
        amount_str = request.POST.get('amount', '').replace(',', '')
        pay_method = request.POST.get('pay_method')
        cause_id = request.POST.get('cause')

        try:
            amount = int(amount_str)
            if amount < 1000:
                messages.error(request, "حداقل مبلغ ۱۰۰۰ تومان است.")
                return redirect('donation:donate_for_martyr', martyr_id=martyr_id)
        except ValueError:
            messages.error(request, "مبلغ وارد شده معتبر نیست.")
            return redirect('donation:donate_for_martyr', martyr_id=martyr_id)

        if not cause_id:
            messages.error(request, "لطفاً دلیل صدقه را انتخاب کنید.")
            return redirect('donation:donate_for_martyr', martyr_id=martyr_id)

        cause_obj = get_object_or_404(DonationCause, id=cause_id)

        if pay_method == 'wallet':
            success, error_msg, donation = process_wallet_donation(
                request.user,
                amount,
                cause_obj,
                martyr=martyr,
                eternal=None
            )
            if not success:
                messages.error(request, error_msg)
                return redirect('donation:donate_for_martyr', martyr_id=martyr_id)

            messages.success(request, "پرداخت از طریق کیف پول با موفقیت انجام شد.")
            return redirect('donations:donation_detail', pk=donation.id)

        elif pay_method == 'gateway':
            donation = Donation.objects.create(
                amount=amount,
                cause=cause_obj,
                status='PENDING',
                user=request.user,
                martyr=martyr,
                pay_method='gateway'
            )


            # ذخیره آدرس بازگشت در سشن
            request.session["payment_return_url"] = reverse(
                "martyrs:martyr_detail", kwargs={"martyr_id": martyr_id}
            )

            return redirect('donation:fake_gateway', donation_id=donation.id)

        else:
            messages.error(request, "روش پرداخت نامعتبر است.")
            return redirect('donation:donate_for_martyr', martyr_id=martyr_id)

    suggested_amounts = [10000, 25000, 40000, 60000, 100000]
    return render(request, 'donation/donate.html', {
        'causes': causes,
        'martyr': martyr,
        'eternal': None,
        'suggested_amounts': suggested_amounts,
    })

















from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.csrf import csrf_exempt
from donation.models import Donation
from django.contrib import messages


from notification.services import confirm_notification_coupon_purchase



@csrf_exempt
@login_required
def payment_callback(request, donation_id):
    donation = get_object_or_404(Donation, id=donation_id)
    status = request.GET.get('status', 'FAILED')
    ref_id = request.GET.get('ref_id', '')

    donation.status = 'SUCCESS' if status == 'OK' else 'FAILED'
    donation.ref_id = ref_id
    donation.save()

    # برای خرید کوپن، تأیید خرید را انجام بده (صرفاً اگر موفق بوده)
    if donation.status == 'SUCCESS':
        if donation.cause and donation.cause.title == "خرید کوپن ارسال پیام":
            confirm_notification_coupon_purchase(donation, payment_method="bank")

    return redirect('donation:donation_detail', pk=donation.id)







@csrf_exempt
def delete_payment_callback(request, donation_id):
    print("Donatean\CALLBACK CALLED")
    print("status:", request.GET.get('status'))
    print("ref_id:", request.GET.get('ref_id'))

    status = request.GET.get('status')
    ref_id = request.GET.get('ref_id')
    gateway_response = dict(request.GET.lists())

    donation = get_object_or_404(Donation, id=donation_id)

    if status == 'OK':
        donation.status = 'SUCCESS'
        donation.save()
        messages.success(request, "پرداخت با موفقیت انجام شد.")
        # می‌توانید اینجا بقیه منطق کوپن یا نوتیفیکیشن‌ها رو هم اضافه کنید

    else:
        donation.status = 'FAILED'
        donation.save()
        messages.error(request, "پرداخت ناموفق بود.")

    # همیشه به صفحه جزییات پرداخت ریدایرکت کن
    return redirect('donations:donation_detail', pk=donation.id)


@login_required
def fake_bank_gateway(request, donation_id):
    donation = get_object_or_404(Donation, id=donation_id)

    if request.method == 'POST':
        status = request.POST.get('status')
        ref_id = request.POST.get('ref_id', 'FAKE123')

        callback_url = reverse(
            'donation:payment_callback',
            args=[donation.id]
        )
        # اگر GET parameters لازم داری:
        callback_url += f"?status={'OK' if status == 'success' else 'FAILED'}&ref_id={ref_id}"

        return redirect(callback_url)

    return render(request, 'donation/fake_gateway.html', {'donation': donation})








@login_required
def delete_fake_bank_gateway(request, donation_id):
    donation = get_object_or_404(Donation, id=donation_id)

    if request.method == 'POST':
        status = request.POST.get('status')
        if status == 'success':
            donation.status = 'SUCCESS'
            messages.success(request, "پرداخت با موفقیت انجام شد.")
        else:
            donation.status = 'FAILED'
            messages.error(request, "پرداخت ناموفق بود.")
        donation.save()

        return redirect('donations:donation_detail', pk=donation.id)

    return render(request, 'donation/fake_gateway.html', {'donation': donation})












def fake_wallet_charge_gateway(request, wallet_tx_id):
    wallet_tx = get_object_or_404(WalletTransaction, id=wallet_tx_id)
    return redirect(f'/wallet/charge/callback/{wallet_tx.id}/?status=OK&ref_id=FAKE12345')



# donation/views.py
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect, get_object_or_404
from .models import DonationCause
from .forms import DonationCauseForm

@staff_member_required
def donationcause_list(request):
    causes = DonationCause.objects.all()
    return render(request, 'donation/donationcause_list.html', {'causes': causes})

@staff_member_required
def donationcause_create(request):
    if request.method == 'POST':
        form = DonationCauseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('donation:donationcause_list')
    else:
        form = DonationCauseForm()
    return render(request, 'donation/donationcause_form.html', {'form': form})

@staff_member_required
def donationcause_edit(request, pk):
    cause = get_object_or_404(DonationCause, pk=pk)
    if request.method == 'POST':
        form = DonationCauseForm(request.POST, instance=cause)
        if form.is_valid():
            form.save()
            return redirect('donation:donationcause_list')
    else:
        form = DonationCauseForm(instance=cause)
    return render(request, 'donation/donationcause_form.html', {'form': form})

@staff_member_required
def donationcause_delete(request, pk):
    cause = get_object_or_404(DonationCause, pk=pk)
    if request.method == 'POST':
        cause.delete()
        return redirect('donation:donationcause_list')
    return render(request, 'donation/donationcause_confirm_delete.html', {'cause': cause})




from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.db.models import Q
from donation.models import Donation
from wallet.models import WalletTransaction



from django.contrib.auth.decorators import user_passes_test, login_required
from django.shortcuts import render
from django.utils import timezone
from wallet.models import WalletTransaction
from donation.models import Donation, DonationCause
from accounts.models import User

@login_required
@user_passes_test(lambda u: u.is_superuser)
def donation_wallet_report(request):
    donations = Donation.objects.all()
    wallet_transactions = WalletTransaction.objects.all()
    
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    donation_status = request.GET.get("donation_status")
    transaction_status = request.GET.get("transaction_status")
    cause_id = request.GET.get("cause")
    user_id = request.GET.get("user")

    # فیلتر تاریخ
    if start_date:
        donations = donations.filter(created_at__date__gte=start_date)
        wallet_transactions = wallet_transactions.filter(created_at__date__gte=start_date)
    if end_date:
        donations = donations.filter(created_at__date__lte=end_date)
        wallet_transactions = wallet_transactions.filter(created_at__date__lte=end_date)

    # فیلتر وضعیت
    if donation_status:
        donations = donations.filter(status=donation_status)
    if transaction_status:
        wallet_transactions = wallet_transactions.filter(status=transaction_status)
    
    # فیلتر علت
    if cause_id:
        donations = donations.filter(cause_id=cause_id)
        wallet_transactions = wallet_transactions.filter(cause_id=cause_id)
    
    # فیلتر کاربر
    if user_id:
        donations = donations.filter(user_id=user_id)
        wallet_transactions = wallet_transactions.filter(wallet__user_id=user_id)

    causes = DonationCause.objects.all()
    users = User.objects.all()

    # Excel export
    if request.GET.get("export") == "excel":
        return export_to_excel(donations, wallet_transactions)

    return render(request, 'donation/report.html', {
        'donations': donations,
        'wallet_transactions': wallet_transactions,
        'causes': causes,
        'users': users,
    })


import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone

def export_to_excel(donations, wallet_transactions):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "صدقات"

    headers1 = ["مبلغ", "وضعیت", "علت", "کاربر", "شهید", "جاودانه", "تاریخ"]
    ws1.append(headers1)

    for d in donations:
        ws1.append([
            d.amount,
            d.get_status_display(),
            str(d.cause or ""),
            str(d.user or ""),
            str(d.martyr or ""),
            str(d.eternal or ""),
            timezone.localtime(d.created_at).strftime("%Y-%m-%d %H:%M")
        ])

    # Sheet 2: Wallet Transactions
    ws2 = wb.create_sheet(title="تراکنش‌های کیف پول")
    headers2 = ["کاربر", "نوع تراکنش", "مبلغ", "وضعیت", "علت", "تاریخ"]
    ws2.append(headers2)

    for tx in wallet_transactions:
        ws2.append([
            str(tx.wallet.user or ""),
            tx.get_transaction_type_display(),
            tx.amount,
            tx.get_status_display(),
            str(tx.cause or ""),
            timezone.localtime(tx.created_at).strftime("%Y-%m-%d %H:%M")
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"donation_wallet_report_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response





from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from .models import Donation



from django.shortcuts import render, get_object_or_404
from .models import Donation





from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from django.urls import reverse
from donation.models import Donation

@login_required
def donation_detail(request, pk):
    donation = get_object_or_404(Donation, id=pk)

    is_system_payment = False
    system_link = None
    system_failed_link = None

    if donation.cause and donation.cause.title == "خرید کوپن ارسال پیام":
        is_system_payment = True
        system_link = reverse('notification:notifications_list')
        system_failed_link = reverse('notification:buy_coupon')

    context = {
        "donation": donation,
        "is_system_payment": is_system_payment,
        "system_link": system_link,
        "system_failed_link": system_failed_link,
    }
    return render(request, "donation/donation_detail.html", context)








# donation/views.py

from django.views.generic import ListView
from django.db.models import Sum, Q
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import Donation









import jdatetime
from datetime import datetime, time

class UserDonationListView(LoginRequiredMixin, ListView):
    model = Donation
    template_name = 'donation/user_donation_list.html'
    context_object_name = 'donations'
    paginate_by = 20

    def get_queryset(self):
        user = self.request.user
        qs = Donation.objects.filter(user=user)

        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')

        start_datetime = None
        end_datetime = None

        if start_date_str:
            try:
                y, m, d = map(int, start_date_str.replace('-', '/').split('/'))
                gregorian_date = jdatetime.date(y, m, d).togregorian()
                start_datetime = datetime.combine(gregorian_date, time.min)
            except:
                start_datetime = None

        if end_date_str:
            try:
                y, m, d = map(int, end_date_str.replace('-', '/').split('/'))
                gregorian_date = jdatetime.date(y, m, d).togregorian()
                end_datetime = datetime.combine(gregorian_date, time.max)
            except:
                end_datetime = None

        # حالا تصمیم می‌گیریم روی چه بازه‌ای فیلتر کنیم
        if start_datetime and end_datetime:
            qs = qs.filter(created_at__range=(start_datetime, end_datetime))
        elif start_datetime:
            qs = qs.filter(created_at__gte=start_datetime)
        elif end_datetime:
            qs = qs.filter(created_at__lte=end_datetime)
        # اگر هیچ‌کدام نبود → بدون فیلتر تاریخ

        status = self.request.GET.get('status')
        if status in ['SUCCESS', 'FAILED', 'PENDING']:
            qs = qs.filter(status=status)

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        qs = self.get_queryset()
        total = qs.aggregate(total_amount=Sum('amount'))['total_amount'] or 0
        context['total_amount'] = total
        return context











